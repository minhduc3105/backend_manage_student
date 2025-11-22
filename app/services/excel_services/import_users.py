from fastapi import UploadFile
from sqlalchemy.orm import Session
from openpyxl import load_workbook  # type: ignore
from io import BytesIO

from .. import service_helper
from app.schemas.user_schema import UserCreate
from app.schemas.user_role_schema import UserRoleCreate
from app.schemas.student_schema import StudentCreate
from app.schemas.parent_schema import ParentCreate

from app.crud.user_crud import create_user
from app.crud.user_role_crud import create_user_role
from app.crud.student_crud import create_student, get_student
from app.crud.parent_crud import create_parent


def import_users(file: UploadFile, db: Session):
    try:
        contents = file.file.read()
        workbook = load_workbook(filename=BytesIO(contents), data_only=True)

        # --- Đọc sheet Student ---
        if "Student" not in workbook.sheetnames:
            raise ValueError("Excel file must contain a 'Student' sheet")
        ws_student = workbook["Student"]
        student_rows = []
        for row in ws_student.iter_rows(min_row=2, values_only=True):
            row = row[1:]  # bỏ cột A (STT)
            clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
            if clean_row[0] and "@" in clean_row[0]:
                student_rows.append(clean_row)

        email_to_student_id = import_students(db, student_rows)

        # --- Đọc sheet Parent ---
        if "Parent" not in workbook.sheetnames:
            raise ValueError("Excel file must contain a 'Parent' sheet")
        ws_parent = workbook["Parent"]
        parent_rows = []
        for row in ws_parent.iter_rows(min_row=2, values_only=True):
            row = row[1:]
            clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
            if clean_row[0] and "@" in clean_row[0]:
                parent_rows.append(clean_row)

        email_to_parent_id = import_parents(db, parent_rows, email_to_student_id)

        return {
            "students": email_to_student_id,
            "parents": email_to_parent_id,
        }

    except Exception as e:
        raise RuntimeError(f"Import failed: {str(e)}")


def import_students(db: Session, student_rows: list) -> dict:
    """
    Import học sinh từ sheet (đã cắt cột A).
    Trả về dict email -> student_id.
    Cột: 0:B(email), 1:C(full_name), 2:D(dob), 3:E(gender), 4:F(phone).
    """
    email_to_student_id = {}

    for row in student_rows:
        if not row:
            continue
        row = list(row) + [""] * (5 - len(row))

        email        = (row[0] or "").strip().lower()
        full_name    = (row[1] or "").strip()
        dob          = row[2]
        gender       = (row[3] or "").strip()
        phone_number = (row[4] or "").strip()

        if not email or "@" not in email:
            continue

        username = email.split("@")[0]
        raw_password = username + "123"
        dob_parsed = service_helper.parse_date_safe(dob)

        # Tạo user
        user_in = UserCreate(
            username=username,
            full_name=full_name,
            email=email,
            password=raw_password,
            date_of_birth=dob_parsed,
            gender=gender,
            phone_number=phone_number,
        )
        db_user = create_user(db, user_in)

        # Gán role student
        create_user_role(db, UserRoleCreate(user_id=db_user.user_id, role_name="student"))

        # Tạo student
        student_in = StudentCreate(user_id=db_user.user_id)
        db_student = create_student(db, student_in)

        email_to_student_id[email] = db_student.user_id

    return email_to_student_id


def import_parents(db: Session, parent_rows: list, email_to_student_id: dict) -> dict:
    """
    Import phụ huynh từ sheet (đã cắt cột A).
    Trả về dict email -> parent_id.
    Đồng thời gán parent_id cho student nếu có child_email.
    Cột: 0:B(email), 1:C(full_name), 2:D(dob), 3:E(gender),
         4:F(phone), 5:G(...), 6:H(...), 7:I(child_email)
    """
    email_to_parent_id = {}

    for row in parent_rows:
        if not row:
            continue
        row = list(row) + [""] * (8 - len(row))

        email        = (row[0] or "").strip().lower()
        full_name    = (row[1] or "").strip()
        dob          = row[2]
        gender       = (row[3] or "").strip()
        phone_number = (row[4] or "").strip()
        child_email  = (row[7] or "").strip().lower()

        if not email or "@" not in email:
            continue

        username = email.split("@")[0]
        raw_password = username + "123"
        dob_parsed = service_helper.parse_date_safe(dob)

        # Tạo user phụ huynh
        user_in = UserCreate(
            username=username,
            full_name=full_name,
            email=email,
            password=raw_password,
            date_of_birth=dob_parsed,
            gender=gender,
            phone_number=phone_number,
        )
        db_user = create_user(db, user_in)

        # Gán role parent
        create_user_role(db, UserRoleCreate(user_id=db_user.user_id, role_name="parent"))

        # Tạo parent
        parent_in = ParentCreate(user_id=db_user.user_id)
        db_parent = create_parent(db, parent_in)
        email_to_parent_id[email] = db_parent.user_id

        # Nếu có child_email hợp lệ -> update student.parent_id
        if child_email and child_email in email_to_student_id:
            student_id = email_to_student_id[child_email]
            student = get_student(db, student_id)
            if student:
                student.parent_id = db_parent.user_id
                db.add(student)
                db.commit()
                db.refresh(student)

    return email_to_parent_id
