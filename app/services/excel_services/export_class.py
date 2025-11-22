from sqlalchemy.orm import Session
from openpyxl import Workbook # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from fastapi.responses import StreamingResponse
from io import BytesIO

from app.models.class_model import Class
from app.models.teacher_model import Teacher
from app.models.student_model import Student
from app.models.user_model import User
from app.models.enrollment_model import Enrollment, EnrollmentStatus

def export_class(db: Session, class_id: int):
    """
    Xuất dữ liệu lớp học ra Excel:
    - B1: class_name
    - D1: teacher full_name
    - B3-D3: header
    - Dữ liệu student lấy từ bảng Enrollment
    """

    # --- Truy vấn thông tin class ---
    db_class = db.query(Class).filter(Class.class_id == class_id).first()
    if not db_class:
        raise ValueError(f"Class id={class_id} not found")

    # --- Truy vấn teacher ---
    teacher_name = ""
    if db_class.teacher_user_id:
        teacher_user = (
            db.query(User)
            .join(Teacher, Teacher.user_id == User.user_id)
            .filter(Teacher.user_id == db_class.teacher_user_id)
            .first()
        )
        teacher_name = teacher_user.full_name if teacher_user else ""

    # --- Truy vấn students qua Enrollment ---
    students = (
        db.query(Student.user_id, User.full_name, User.date_of_birth, User.email, User.phone_number, User.gender)
        .join(User, User.user_id == Student.user_id)
        .join(Enrollment, Enrollment.student_user_id == Student.user_id)
        .filter(
            Enrollment.class_id == class_id,
            Enrollment.enrollment_status == EnrollmentStatus.active
        )
        .all()
    )

    # --- Tạo file Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header thông tin class
    ws["A1"] = f"Class: {db_class.class_name}"
    ws["C1"] = f"Teacher: {teacher_name}"

    # Header bảng students
    ws["A3"] = "STT"
    ws["B3"] = "Student ID"
    ws["C3"] = "Full name"
    ws["D3"] = "Date of birth"
    ws["E3"] = "Email"
    ws["F3"] = "Phone number"
    ws["G3"] = "Gender"

    # Ghi dữ liệu students
    for idx, st in enumerate(students, start=1):
        ws.cell(row=3 + idx, column=1, value=idx)
        ws.cell(row=3 + idx, column=2, value=st.user_id)
        ws.cell(row=3 + idx, column=3, value=st.full_name)
        
        # Format date to dd/mm/yyyy
        dob_formatted = st.date_of_birth.strftime("%d/%m/%Y") if st.date_of_birth else ""
        ws.cell(row=3 + idx, column=4, value=dob_formatted)
        
        ws.cell(row=3 + idx, column=5, value=st.email)
        ws.cell(row=3 + idx, column=6, value=st.phone_number)
        ws.cell(row=3 + idx, column=7, value=st.gender.value if st.gender else "")

    # Auto-adjust column width to fit content
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_length = 0
        for cell in ws[col_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Export to response
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"class_{class_id}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )